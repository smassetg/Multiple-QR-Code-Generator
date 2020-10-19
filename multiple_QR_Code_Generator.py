# Import required packages for program to work.
import pyqrcode
from pyqrcode import QRCode
import openpyxl as op

# Get string to make the QR code from excel sheet
website_list = 'XXXXXX' # The name of the website list - enter the path of your excel file here
wb1 = op.load_workbook(website_list) # Load the website work book with workbook criteria below
ws1 = wb1.active
mr = ws1.max_row
mc = ws1.max_column

for i in range (1, mr+1):
    for j in range (1, mc+1):
        #This cycles through the names in the excel sheet, and saves in the 'new_name' variable
        string = ws1.cell(row = i, column = j).value
        print("i =", i) # These lines shows how the for loop is working and can be commented out / deleted
        print("j =", j)
        print("string =", string)

        # Generate QR code
        url = pyqrcode.create(string)

        # Create and save the png file named after the string
        url.svg(string + '.svg', scale = 8)
